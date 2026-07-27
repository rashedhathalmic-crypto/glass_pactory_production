from pathlib import Path
from openpyxl import load_workbook
import pytest
from cnc_nc_generator.cli import main
from cnc_nc_generator.native_nc import NCParameters, SUPPORTED_PROFILES, _geometry, generate_nc

WORKBOOK=Path(__file__).parents[1]/'Glass Factory Programs - Template - small fixture.xlsx'
ROWS={'129-122-03-210':3,'129-122-03-102':5,'129-122-03-211':7}
@pytest.mark.parametrize('profile',SUPPORTED_PROFILES)
def test_all_profile_coordinates_match_workbook(profile):
    ws=load_workbook(WORKBOOK,data_only=True)['All parts parameters']; row=ROWS[profile]
    expected=[[ws.cell(row,2+group*4+cut).value for cut in range(4)] for group in range(17)]
    actual=_geometry(NCParameters(profile))
    for a,e in zip(actual,expected): assert a==pytest.approx(e,abs=1e-10)

def test_complete_default_program_matches_workbook():
    ws=load_workbook(WORKBOOK,data_only=True)['Part program']
    expected='\n'.join(str(ws.cell(row,1).value or '') for row in range(1,76))+'\n'
    assert generate_nc(NCParameters('129-122-03-211'))==expected

def test_cli_does_not_need_excel(tmp_path,monkeypatch):
    monkeypatch.chdir(tmp_path); output=tmp_path/'part.nc'
    assert main(['129-122-03-102','-o',str(output)])==0
    assert output.read_text().startswith('%\nO0001\n')
